import csv
import io
import logging
import platform
import socket
import ssl
import subprocess
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
from urllib.request import Request as UrlRequest, urlopen

import psutil
from fastapi import Depends, FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, PlainTextResponse, StreamingResponse
from pydantic import BaseModel, ValidationError
from sqlalchemy import String, cast, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from . import models, schemas
from .apple_calendar import apple_calendar_config, apple_calendar_status, list_apple_calendars, save_apple_calendar_config, sync_apple_calendar
from .calendar_reminders import auto_complete_past_calendar_reminders, backfill_calendar_reminders
from .database import Base, DATA_DIR, engine, get_db
from .gmail_bills import connect_gmail, gmail_status, scan_gmail_for_bills
from .google_calendar import connect_google_calendar, google_calendar_status, sync_google_calendar
from .google_oauth_config import google_oauth_config_status, save_google_oauth_config
from .migrations import ensure_schema

try:
    import dns.resolver
except ImportError:  # pragma: no cover - optional runtime support
    dns = None

try:
    import winreg
except ImportError:  # pragma: no cover - non-Windows runtime support
    winreg = None

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s %(message)s")
logger = logging.getLogger("ai_home_dashboard")

Base.metadata.create_all(bind=engine)
ensure_schema(engine)

app = FastAPI(title="AI Home Dashboard API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5174", "http://127.0.0.1:5174"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type"],
)


@app.exception_handler(HTTPException)
async def http_exception_handler(_: Request, exc: HTTPException):
    return JSONResponse(status_code=exc.status_code, content={"error": {"message": exc.detail}})


@app.exception_handler(ValidationError)
async def validation_exception_handler(_: Request, exc: ValidationError):
    return JSONResponse(status_code=422, content={"error": {"message": "Validation failed", "details": exc.errors()}})


@app.get("/health")
def health():
    return {"status": "ok", "time": datetime.now().isoformat(timespec="seconds")}


def directory_size(path):
    total = 0
    for item in path.rglob("*"):
        if item.is_file():
            try:
                total += item.stat().st_size
            except OSError:
                continue
    return total


def storage_metrics():
    drives = []
    seen_mounts = set()
    for partition in psutil.disk_partitions(all=False):
        mountpoint = partition.mountpoint
        if not mountpoint or mountpoint in seen_mounts:
            continue
        seen_mounts.add(mountpoint)
        try:
            usage = psutil.disk_usage(mountpoint)
        except OSError:
            continue
        drives.append(
            {
                "device": partition.device,
                "mountpoint": mountpoint,
                "filesystem": partition.fstype,
                "total_gb": round(usage.total / (1024**3), 2),
                "used_gb": round(usage.used / (1024**3), 2),
                "free_gb": round(usage.free / (1024**3), 2),
                "percent": round(usage.percent, 1),
            }
        )
    return {
        "drives": drives,
        "app_data": {
            "path": str(DATA_DIR),
            "size_mb": round(directory_size(DATA_DIR) / (1024**2), 2),
        },
    }


def registry_value(key, name: str) -> str | None:
    try:
        value, _ = winreg.QueryValueEx(key, name)
    except OSError:
        return None
    if value in (None, ""):
        return None
    return str(value)


def installed_windows_apps() -> list[dict[str, Any]]:
    if winreg is None:
        return []
    roots = [
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall", "Machine"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall", "Machine 32-bit"),
        (winreg.HKEY_CURRENT_USER, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall", "Current User"),
    ]
    apps: dict[tuple[str, str, str], dict[str, Any]] = {}
    for hive, path, scope in roots:
        try:
            parent = winreg.OpenKey(hive, path)
        except OSError:
            continue
        with parent:
            for index in range(winreg.QueryInfoKey(parent)[0]):
                try:
                    subkey_name = winreg.EnumKey(parent, index)
                    subkey = winreg.OpenKey(parent, subkey_name)
                except OSError:
                    continue
                with subkey:
                    name = registry_value(subkey, "DisplayName")
                    if not name:
                        continue
                    system_component = registry_value(subkey, "SystemComponent")
                    release_type = registry_value(subkey, "ReleaseType")
                    parent_key = registry_value(subkey, "ParentKeyName")
                    if system_component == "1" or release_type or parent_key:
                        continue
                    publisher = registry_value(subkey, "Publisher") or ""
                    version = registry_value(subkey, "DisplayVersion") or ""
                    install_date = registry_value(subkey, "InstallDate") or ""
                    install_location = registry_value(subkey, "InstallLocation") or ""
                    app_key = (name.casefold(), publisher.casefold(), version.casefold())
                    apps[app_key] = {
                        "name": name,
                        "publisher": publisher,
                        "version": version,
                        "install_date": install_date,
                        "install_location": install_location,
                        "scope": scope,
                        "registry_key": subkey_name,
                    }
    return sorted(apps.values(), key=lambda item: (item["name"].casefold(), item["publisher"].casefold()))


@app.get("/api/system/health")
def system_health():
    memory = psutil.virtual_memory()
    cpu_percent = psutil.cpu_percent(interval=0.1)
    storage = storage_metrics()
    status = "healthy"
    if cpu_percent >= 90 or memory.percent >= 90:
        status = "critical"
    elif cpu_percent >= 75 or memory.percent >= 80:
        status = "watch"
    if any(drive["percent"] >= 95 for drive in storage["drives"]):
        status = "critical"
    elif status == "healthy" and any(drive["percent"] >= 85 for drive in storage["drives"]):
        status = "watch"
    return {
        "status": status,
        "cpu_percent": round(cpu_percent, 1),
        "memory_percent": round(memory.percent, 1),
        "memory_used_gb": round(memory.used / (1024**3), 2),
        "memory_total_gb": round(memory.total / (1024**3), 2),
        "processor_count": psutil.cpu_count(logical=True),
        "host": platform.node(),
        "storage": storage,
        "checked_at": datetime.now().isoformat(timespec="seconds"),
    }


MODEL_CONFIG: dict[str, dict[str, Any]] = {
    "calendar": {"model": models.CalendarItem, "create": schemas.CalendarCreate, "update": schemas.CalendarUpdate, "out": schemas.CalendarOut, "date": "date"},
    "reminders": {"model": models.Reminder, "create": schemas.ReminderCreate, "update": schemas.ReminderUpdate, "out": schemas.ReminderOut, "date": "due_date"},
    "bills": {"model": models.Bill, "create": schemas.BillCreate, "update": schemas.BillUpdate, "out": schemas.BillOut, "date": "due_date"},
    "chores": {"model": models.Chore, "create": schemas.ChoreCreate, "update": schemas.ChoreUpdate, "out": schemas.ChoreOut, "date": "due_date"},
    "notes": {"model": models.Note, "create": schemas.NoteCreate, "update": schemas.NoteUpdate, "out": schemas.NoteOut, "date": "updated_at"},
    "assets": {"model": models.Asset, "create": schemas.AssetCreate, "update": schemas.AssetUpdate, "out": schemas.AssetOut, "date": "updated_at"},
    "speed-tests": {"model": models.SpeedTestResult, "create": schemas.SpeedTestResultCreate, "update": schemas.SpeedTestResultUpdate, "out": schemas.SpeedTestResultOut, "date": "tested_at"},
    "security-records": {"model": models.SecurityRecord, "create": schemas.SecurityRecordCreate, "update": schemas.SecurityRecordUpdate, "out": schemas.SecurityRecordOut, "date": "next_review_date"},
    "links": {"model": models.QuickLink, "create": schemas.QuickLinkCreate, "update": schemas.QuickLinkUpdate, "out": schemas.QuickLinkOut, "date": "updated_at"},
    "projects": {"model": models.Project, "create": schemas.ProjectCreate, "update": schemas.ProjectUpdate, "out": schemas.ProjectOut, "date": "due_date"},
    "project-tasks": {"model": models.ProjectTask, "create": schemas.ProjectTaskCreate, "update": schemas.ProjectTaskUpdate, "out": schemas.ProjectTaskOut, "date": "due_date"},
    "inbox": {"model": models.InboxItem, "create": schemas.InboxCreate, "update": schemas.InboxUpdate, "out": schemas.InboxOut, "date": "created_at"},
    "daily-plans": {"model": models.DailyPlan, "create": schemas.DailyPlanCreate, "update": schemas.DailyPlanUpdate, "out": schemas.DailyPlanOut, "date": "date"},
}


def serialize(items: list[Any], schema: type[BaseModel]):
    return [schema.model_validate(item).model_dump() for item in items]


def normalize_header(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def read_delimited_bill_rows(content: bytes, suffix: str) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel_tab if suffix == ".txt" else csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    return rows_to_bill_dicts(rows)


def read_xlsx_bill_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise HTTPException(status_code=500, detail="XLSX import requires openpyxl. Install backend requirements first.") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    return rows_to_bill_dicts(rows)


def rows_to_bill_dicts(rows: list[list[Any]]) -> list[dict[str, Any]]:
    cleaned = [[cell for cell in row] for row in rows if any(str(cell or "").strip() for cell in row)]
    if not cleaned:
        return []
    first_row = [normalize_header(cell) for cell in cleaned[0]]
    known_headers = {
        "name", "bill", "vendor", "company", "merchant", "payee", "service", "subscription",
        "amount", "cost", "price", "total", "duedate", "due", "date", "billingcycle", "cycle",
        "frequency", "category", "autopay", "auto", "url", "link", "status", "notes", "note",
        "description", "subject",
    }
    has_header = len(set(first_row) & known_headers) >= 2
    headers = first_row if has_header else [
        "name", "amount", "duedate", "billingcycle", "category", "autopay", "url", "status", "notes"
    ]
    data_rows = cleaned[1:] if has_header else cleaned
    records: list[dict[str, Any]] = []
    for row in data_rows:
        record = {}
        for index, header in enumerate(headers):
            if index < len(row):
                record[header] = row[index]
        records.append(record)
    return records


ARTICLE_IMPORT_HEADERS = ["title", "category", "tags", "pinned", "body"]
LINK_IMPORT_HEADERS = ["name", "url", "category", "environment", "tags", "favorite", "body"]
NOTE_CATEGORIES = {"note", "runbook", "troubleshooting", "script", "codex_prompt", "decision", "reference", "how_to", "other"}
LINK_ENVIRONMENTS = {"local", "personal", "work", "public", "other"}


def read_delimited_article_rows(content: bytes) -> list[list[Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect)]


def read_xlsx_article_rows(content: bytes) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise HTTPException(status_code=500, detail="XLSX import requires openpyxl. Install backend requirements first.") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]


def rows_to_article_dicts(rows: list[list[Any]]) -> list[dict[str, Any]]:
    cleaned = [[cell for cell in row] for row in rows if any(str(cell or "").strip() for cell in row)]
    if not cleaned:
        return []
    headers = [normalize_header(cell) for cell in cleaned[0][:5]]
    if headers != ARTICLE_IMPORT_HEADERS:
        raise HTTPException(status_code=422, detail="Article import requires columns in this order: Title, Category, Tags, Pinned, Body")
    records: list[dict[str, Any]] = []
    for row in cleaned[1:]:
        records.append({
            "title": row[0] if len(row) > 0 else "",
            "category": row[1] if len(row) > 1 else "",
            "tags": row[2] if len(row) > 2 else "",
            "pinned": row[3] if len(row) > 3 else "",
            "body": row[4] if len(row) > 4 else "",
        })
    return records


def rows_to_link_dicts(rows: list[list[Any]]) -> list[dict[str, Any]]:
    cleaned = [[cell for cell in row] for row in rows if any(str(cell or "").strip() for cell in row)]
    if not cleaned:
        return []
    headers = [normalize_header(cell) for cell in cleaned[0][:7]]
    if headers != LINK_IMPORT_HEADERS:
        raise HTTPException(status_code=422, detail="Link import requires columns in this order: Name, URL, Category, Environment, Tags, Favorite, Body")
    records: list[dict[str, Any]] = []
    for row in cleaned[1:]:
        records.append({
            "name": row[0] if len(row) > 0 else "",
            "url": row[1] if len(row) > 1 else "",
            "category": row[2] if len(row) > 2 else "",
            "environment": row[3] if len(row) > 3 else "",
            "tags": row[4] if len(row) > 4 else "",
            "favorite": row[5] if len(row) > 5 else "",
            "body": row[6] if len(row) > 6 else "",
        })
    return records


def parse_true_false(value: Any, label: str = "Pinned") -> bool:
    if value is None:
        return False
    raw = str(value).strip().upper()
    if raw == "":
        return False
    if raw == "TRUE":
        return True
    if raw == "FALSE":
        return False
    raise ValueError(f"{label} must be TRUE or FALSE, got {value!r}")


def normalize_note_category(value: Any) -> str:
    category = str(value or "").strip().lower()
    if category not in NOTE_CATEGORIES:
        raise ValueError("Category must match an accepted Knowledge Base category")
    return category


def normalize_link_environment(value: Any) -> str:
    environment = str(value or "personal").strip().lower()
    if not environment:
        return "personal"
    if environment not in LINK_ENVIRONMENTS:
        raise ValueError(f"Environment must be one of {', '.join(sorted(LINK_ENVIRONMENTS))}")
    return environment


def normalize_article_import_row(row: dict[str, Any]) -> dict[str, Any]:
    title = str(row.get("title") or "").strip()
    if not title:
        raise ValueError("Title is required")
    payload = {
        "title": title[:160],
        "note_type": normalize_note_category(row.get("category")),
        "tags": str(row.get("tags") or "").strip()[:300] or None,
        "pinned": parse_true_false(row.get("pinned")),
        "body": str(row.get("body") or ""),
    }
    return schemas.NoteCreate.model_validate(payload).model_dump()


def normalize_link_import_row(row: dict[str, Any]) -> dict[str, Any]:
    name = str(row.get("name") or "").strip()
    url = str(row.get("url") or "").strip()
    if not name:
        raise ValueError("Name is required")
    if not url:
        raise ValueError("URL is required")
    payload = {
        "name": name[:120],
        "url": url,
        "category": str(row.get("category") or "").strip()[:80] or None,
        "environment": normalize_link_environment(row.get("environment")),
        "tags": str(row.get("tags") or "").strip()[:300] or None,
        "favorite": parse_true_false(row.get("favorite"), "Favorite"),
        "notes": str(row.get("body") or ""),
    }
    return schemas.QuickLinkCreate.model_validate(payload).model_dump()


def duplicate_text_key(value: Any, *, casefold: bool = True) -> str:
    text = str(value or "").strip()
    return text.casefold() if casefold else text


def article_import_duplicate_key(item: dict[str, Any]) -> tuple[str, str, str]:
    return (
        duplicate_text_key(item.get("title")),
        duplicate_text_key(item.get("note_type")),
        duplicate_text_key(item.get("body"), casefold=False),
    )


def link_import_duplicate_key(item: dict[str, Any]) -> tuple[str, str]:
    return (
        duplicate_text_key(item.get("name")),
        duplicate_text_key(item.get("url")),
    )


def parse_bool(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "autopay", "auto", "automatic"}


def parse_amount(value: Any) -> float:
    clean = str(value or "0").replace("$", "").replace(",", "").strip()
    return float(clean or 0)


def parse_date_value(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return date.fromisoformat(raw[:10]).isoformat()
    except ValueError:
        return None


def field_value(row: dict[str, Any], aliases: list[str]) -> Any:
    for alias in aliases:
        key = normalize_header(alias)
        if key in row and str(row[key] or "").strip():
            return row[key]
    return None


def normalize_cycle(value: Any) -> str:
    raw = str(value or "monthly").strip().lower().replace("_", "-")
    aliases = {"annual": "yearly", "annually": "yearly", "year": "yearly", "month": "monthly", "quarter": "quarterly", "week": "weekly", "once": "one-time", "onetime": "one-time"}
    return aliases.get(raw, raw if raw in {"weekly", "monthly", "quarterly", "yearly", "one-time"} else "monthly")


def normalize_bill_import_row(row: dict[str, Any], row_number: int) -> dict[str, Any] | None:
    name = field_value(row, ["name", "bill", "vendor", "company", "merchant", "payee", "service", "subscription", "subject"])
    due_date = parse_date_value(field_value(row, ["due_date", "due date", "duedate", "due", "date"]))
    if not name or not due_date:
        return None
    notes = field_value(row, ["notes", "note", "description", "memo"])
    subject = field_value(row, ["subject"])
    payload = {
        "name": str(name).strip()[:160],
        "amount": parse_amount(field_value(row, ["amount", "cost", "price", "total", "monthly amount"])),
        "due_date": due_date,
        "billing_cycle": normalize_cycle(field_value(row, ["billing_cycle", "billing cycle", "cycle", "frequency"])),
        "category": field_value(row, ["category", "type"]),
        "autopay": parse_bool(field_value(row, ["autopay", "auto pay", "auto", "automatic"])),
        "url": field_value(row, ["url", "link", "website"]),
        "notes": str(notes or subject or f"Imported from bill file row {row_number}").strip(),
        "status": str(field_value(row, ["status"]) or "review").strip().lower(),
        "source": "manual",
    }
    return schemas.BillCreate.model_validate(payload).model_dump()


def now_stamp() -> str:
    return datetime.now().isoformat(timespec="seconds")


def suggest_capture_type(raw_text: str) -> tuple[str, str]:
    text = raw_text.lower()
    if "http://" in text or "https://" in text:
        return "link", "medium"
    if any(word in text for word in ["remind", "todo", "follow up", "due", "call ", "email "]):
        return "reminder", "high" if any(word in text for word in ["urgent", "asap", "today"]) else "medium"
    if any(word in text for word in ["meeting", "appointment", "calendar", "schedule"]):
        return "calendar item", "medium"
    if any(word in text for word in ["codex", "repo", "project", "bug", "feature", "implement", "fix "]):
        return "project task", "medium"
    if any(word in text for word in ["idea", "maybe", "could "]):
        return "idea", "low"
    return "note", "medium"


def first_url(text: str) -> str | None:
    for token in text.split():
        clean = token.strip("()[]{}<>,")
        if clean.startswith(("http://", "https://")):
            return clean
    return None


def is_local_network_url(value: str | None) -> bool:
    if not value:
        return False
    host = urlparse(value).hostname or ""
    if host in {"localhost", "::1"} or host.startswith(("10.", "192.168.", "127.")):
        return True
    parts = host.split(".")
    return len(parts) == 4 and parts[0] == "172" and parts[1].isdigit() and 16 <= int(parts[1]) <= 31


def project_display(project: models.Project) -> dict[str, Any]:
    payload = schemas.ProjectOut.model_validate(project).model_dump()
    payload["status"] = schemas.normalize_project_status(payload["status"])
    payload["local_path"] = payload["local_path"] or payload["codex_workspace_path"]
    payload["repo_url"] = payload["repo_url"] or payload["repository_url"]
    payload["next_action"] = payload["next_action"] or payload["next_step"]
    payload["goal"] = payload["goal"] or payload["description"]
    return payload


def link_display(link: models.QuickLink) -> dict[str, Any]:
    payload = schemas.QuickLinkOut.model_validate(link).model_dump()
    payload["local_network"] = is_local_network_url(link.url)
    return payload


def security_record_display(record: models.SecurityRecord) -> dict[str, Any]:
    return schemas.SecurityRecordOut.model_validate(record).model_dump()


def asset_display(asset: models.Asset) -> dict[str, Any]:
    return schemas.AssetOut.model_validate(asset).model_dump()


def reminder_to_agenda_item(reminder: models.Reminder) -> dict[str, Any]:
    return {
        "id": f"reminder-{reminder.id}",
        "record_id": reminder.id,
        "title": reminder.title,
        "date": reminder.due_date,
        "start_time": reminder.due_time,
        "end_time": None,
        "location": None,
        "notes": reminder.notes,
        "category": "Reminder",
        "source": "reminder",
        "priority": reminder.priority,
        "status": reminder.status,
    }


def calendar_to_agenda_item(item: models.CalendarItem) -> dict[str, Any]:
    payload = schemas.CalendarOut.model_validate(item).model_dump()
    payload["record_id"] = item.id
    return payload


def combined_agenda(db: Session, start_date: date, end_date: date) -> list[dict[str, Any]]:
    start = start_date.isoformat()
    end = end_date.isoformat()
    calendar_items = db.scalars(
        select(models.CalendarItem)
        .where(models.CalendarItem.date.between(start, end))
        .order_by(models.CalendarItem.date, models.CalendarItem.start_time)
    ).all()
    reminders = db.scalars(
        select(models.Reminder)
        .where(
            models.Reminder.status == "open",
            models.Reminder.due_date.is_not(None),
            models.Reminder.due_date.between(start, end),
            or_(models.Reminder.source == "manual", models.Reminder.source.is_(None)),
        )
        .order_by(models.Reminder.due_date, models.Reminder.due_time)
    ).all()
    items = [calendar_to_agenda_item(item) for item in calendar_items]
    items.extend(reminder_to_agenda_item(reminder) for reminder in reminders)
    return sorted(items, key=lambda item: (item["date"] or "9999-99-99", item["start_time"] or "99:99", item["title"]))


def get_or_404(db: Session, model: type, item_id: int):
    item = db.get(model, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Record not found")
    return item


def commit_or_conflict(db: Session):
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        logger.warning("Integrity error: %s", exc)
        raise HTTPException(status_code=409, detail="A similar record already exists") from exc


def next_due(current: str | None, recurrence: str) -> str | None:
    if not current or recurrence == "one-time":
        return current
    base = datetime.strptime(current, "%Y-%m-%d").date()
    if recurrence == "daily":
        return (base + timedelta(days=1)).isoformat()
    if recurrence == "weekly":
        return (base + timedelta(days=7)).isoformat()
    if recurrence == "monthly":
        month = base.month + 1
        year = base.year + (1 if month == 13 else 0)
        month = 1 if month == 13 else month
        day = min(base.day, 28)
        return date(year, month, day).isoformat()
    return current


def clamp_host(value: str) -> str:
    host = value.strip()
    if not host or any(char.isspace() for char in host):
        raise HTTPException(status_code=422, detail="Enter a single hostname or IP address")
    if "://" in host:
        parsed = urlparse(host)
        host = parsed.hostname or ""
    return host.strip("[]")


def latency_samples(url: str, count: int = 3) -> list[float]:
    samples = []
    for _ in range(count):
        start = time.perf_counter()
        with urlopen(UrlRequest(url, headers={"User-Agent": "AIHomeDashboard/0.1"}), timeout=8) as response:
            response.read(1)
        samples.append((time.perf_counter() - start) * 1000)
    return samples


def measure_download_mbps(url: str, byte_limit: int = 5_000_000) -> tuple[float | None, int, float]:
    started = time.perf_counter()
    with urlopen(UrlRequest(url, headers={"User-Agent": "AIHomeDashboard/0.1"}), timeout=20) as response:
        chunk = response.read(byte_limit)
    elapsed = max(time.perf_counter() - started, 0.001)
    bytes_read = len(chunk)
    mbps = round((bytes_read * 8) / elapsed / 1_000_000, 2) if bytes_read else None
    return mbps, bytes_read, elapsed


def measure_upload_mbps(url: str, byte_count: int = 1_000_000) -> tuple[float | None, int, float]:
    payload = b"0" * byte_count
    request = UrlRequest(
        url,
        data=payload,
        headers={"Content-Type": "application/octet-stream", "User-Agent": "AIHomeDashboard/0.1"},
        method="POST",
    )
    started = time.perf_counter()
    with urlopen(request, timeout=20) as response:
        response.read(1024)
    elapsed = max(time.perf_counter() - started, 0.001)
    mbps = round((byte_count * 8) / elapsed / 1_000_000, 2)
    return mbps, byte_count, elapsed


def run_safe_ping(host: str) -> dict[str, Any]:
    safe_host = clamp_host(host)
    packet_count = 10
    command = (
        ["ping", "-n", str(packet_count), "-w", "3000", safe_host]
        if platform.system().lower() == "windows"
        else ["ping", "-c", str(packet_count), "-W", "3", safe_host]
    )
    started = time.perf_counter()
    try:
        completed = subprocess.run(command, capture_output=True, text=True, timeout=15, check=False)
    except (OSError, subprocess.TimeoutExpired) as exc:
        logger.exception("Ping failed for host %s", safe_host)
        raise HTTPException(status_code=502, detail="Ping failed") from exc
    elapsed = round((time.perf_counter() - started) * 1000, 2)
    output = (completed.stdout or completed.stderr or "").strip()
    latency = None
    for token in output.replace("<", "=").replace("ms", " ms").split():
        if token.startswith("time=") or token.startswith("time<") or token.startswith("Average"):
            number = "".join(ch for ch in token.split("=")[-1] if ch.isdigit() or ch == ".")
            if number:
                latency = float(number)
                break
    return {
        "host": safe_host,
        "reachable": completed.returncode == 0,
        "latency_ms": latency if latency is not None else (elapsed if completed.returncode == 0 else None),
        "checked_at": now_stamp(),
        "method": "icmp_ping",
        "packet_count": packet_count,
        "message": output[-500:],
    }


def port_check(host: str, port: int) -> dict[str, Any]:
    safe_host = clamp_host(host)
    started = time.perf_counter()
    try:
        with socket.create_connection((safe_host, port), timeout=4):
            latency = round((time.perf_counter() - started) * 1000, 2)
            return {"host": safe_host, "port": port, "reachable": True, "latency_ms": latency, "checked_at": now_stamp()}
    except OSError as exc:
        logger.exception("Port check failed for host %s port %s", safe_host, port)
        return {"host": safe_host, "port": port, "reachable": False, "latency_ms": None, "checked_at": now_stamp(), "error": "Port check failed"}


def certificate_status(days_remaining: int) -> str:
    if days_remaining < 0:
        return "expired"
    if days_remaining <= 30:
        return "expiring_soon"
    return "healthy"


def register_crud(path: str, cfg: dict[str, Any]):
    model = cfg["model"]
    create_schema = cfg["create"]
    update_schema = cfg["update"]
    out_schema = cfg["out"]
    date_field = cfg["date"]

    @app.get(f"/api/{path}")
    def list_items(db: Session = Depends(get_db), limit: int = Query(200, ge=1, le=500), include_completed: bool = Query(False)):
        if path == "reminders":
            auto_complete_past_calendar_reminders(db, commit=True)
        stmt = select(model).limit(limit)
        if path == "reminders" and not include_completed:
            stmt = stmt.where(model.status != "completed")
        if hasattr(model, date_field):
            stmt = stmt.order_by(getattr(model, date_field).is_(None), getattr(model, date_field), model.id)
        return serialize(db.scalars(stmt).all(), out_schema)

    @app.post(f"/api/{path}", status_code=201)
    def create_item(payload: create_schema, db: Session = Depends(get_db)):  # type: ignore[valid-type]
        data = payload.model_dump()
        if path == "inbox":
            suggested_type, suggested_priority = suggest_capture_type(data["raw_text"])
            data["suggested_type"] = suggested_type
            data["priority"] = data.get("priority") or suggested_priority
            data["status"] = "new"
        if path == "projects":
            data["codex_workspace_path"] = data.get("codex_workspace_path") or data.get("local_path")
            data["repository_url"] = data.get("repository_url") or data.get("repo_url")
            data["next_step"] = data.get("next_step") or data.get("next_action")
            data["description"] = data.get("description") or data.get("goal")
        if path == "speed-tests" and not data.get("tested_at"):
            data.pop("tested_at", None)
        item = model(**data)
        db.add(item)
        commit_or_conflict(db)
        db.refresh(item)
        logger.info("Created %s id=%s", path, item.id)
        return out_schema.model_validate(item)

    @app.get(f"/api/{path}" + "/{item_id}")
    def read_item(item_id: int, db: Session = Depends(get_db)):
        return out_schema.model_validate(get_or_404(db, model, item_id))

    @app.put(f"/api/{path}" + "/{item_id}")
    def update_item(item_id: int, payload: update_schema, db: Session = Depends(get_db)):  # type: ignore[valid-type]
        item = get_or_404(db, model, item_id)
        data = payload.model_dump(exclude_unset=True)
        if path == "projects":
            if data.get("local_path") and not data.get("codex_workspace_path"):
                data["codex_workspace_path"] = data["local_path"]
            if data.get("repo_url") and not data.get("repository_url"):
                data["repository_url"] = data["repo_url"]
            if data.get("next_action") and not data.get("next_step"):
                data["next_step"] = data["next_action"]
            if data.get("goal") and not data.get("description"):
                data["description"] = data["goal"]
        for key, value in data.items():
            setattr(item, key, value)
        commit_or_conflict(db)
        db.refresh(item)
        logger.info("Updated %s id=%s", path, item.id)
        return out_schema.model_validate(item)

    @app.delete(f"/api/{path}" + "/{item_id}", status_code=204)
    def delete_item(item_id: int, db: Session = Depends(get_db)):
        item = get_or_404(db, model, item_id)
        db.delete(item)
        db.commit()
        logger.info("Deleted %s id=%s", path, item_id)
        return None


@app.post("/api/bills/import-file")
async def import_bills_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    filename = file.filename or ""
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".txt", ".xlsx"}:
        raise HTTPException(status_code=422, detail="Upload a .csv, .txt, or .xlsx file")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=422, detail="The selected file is empty")
    if len(content) > 5_000_000:
        raise HTTPException(status_code=413, detail="Bill import files must be 5 MB or smaller")
    rows = read_xlsx_bill_rows(content) if suffix == ".xlsx" else read_delimited_bill_rows(content, suffix)
    imported = 0
    skipped_duplicates = 0
    skipped_unmatched = 0
    skipped_invalid = 0
    for index, row in enumerate(rows, start=2):
        try:
            candidate = normalize_bill_import_row(row, index)
        except (ValueError, ValidationError):
            skipped_invalid += 1
            continue
        if not candidate:
            skipped_unmatched += 1
            continue
        exists = db.scalars(
            select(models.Bill).where(
                (models.Bill.name == candidate["name"]) & (models.Bill.due_date == candidate["due_date"])
            )
        ).first()
        if exists:
            skipped_duplicates += 1
            continue
        db.add(models.Bill(**candidate))
        imported += 1
    commit_or_conflict(db)
    return {
        "filename": filename,
        "rows": len(rows),
        "imported": imported,
        "skipped_duplicates": skipped_duplicates,
        "skipped_unmatched": skipped_unmatched,
        "skipped_invalid": skipped_invalid,
    }


@app.post("/api/notes/import-file")
async def import_notes_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    filename = file.filename or ""
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise HTTPException(status_code=422, detail="Upload a .csv or .xlsx article file")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=422, detail="The selected file is empty")
    if len(content) > 5_000_000:
        raise HTTPException(status_code=413, detail="Article import files must be 5 MB or smaller")
    rows = read_xlsx_article_rows(content) if suffix == ".xlsx" else read_delimited_article_rows(content)
    records = rows_to_article_dicts(rows)
    imported = 0
    skipped_duplicates = 0
    skipped_invalid = 0
    errors: list[str] = []
    seen_keys = {
        article_import_duplicate_key({"title": note.title, "note_type": note.note_type, "body": note.body})
        for note in db.scalars(select(models.Note)).all()
    }
    for index, row in enumerate(records, start=2):
        try:
            candidate = normalize_article_import_row(row)
        except (ValueError, ValidationError) as exc:
            skipped_invalid += 1
            errors.append(f"Row {index}: {exc}")
            continue
        duplicate_key = article_import_duplicate_key(candidate)
        if duplicate_key in seen_keys:
            skipped_duplicates += 1
            continue
        db.add(models.Note(**candidate))
        seen_keys.add(duplicate_key)
        imported += 1
    commit_or_conflict(db)
    result = {
        "filename": filename,
        "rows": len(records),
        "imported": imported,
        "skipped_duplicates": skipped_duplicates,
        "skipped_invalid": skipped_invalid,
        "errors": errors[:10],
        "accepted_categories": sorted(NOTE_CATEGORIES),
    }
    logger.info(
        "Imported notes file=%s rows=%s imported=%s skipped_duplicates=%s skipped_invalid=%s",
        filename,
        result["rows"],
        imported,
        skipped_duplicates,
        skipped_invalid,
    )
    return result


@app.post("/api/links/import-file")
async def import_links_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    filename = file.filename or ""
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise HTTPException(status_code=422, detail="Upload a .csv or .xlsx link file")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=422, detail="The selected file is empty")
    if len(content) > 5_000_000:
        raise HTTPException(status_code=413, detail="Link import files must be 5 MB or smaller")
    rows = read_xlsx_article_rows(content) if suffix == ".xlsx" else read_delimited_article_rows(content)
    records = rows_to_link_dicts(rows)
    imported = 0
    skipped_duplicates = 0
    skipped_invalid = 0
    errors: list[str] = []
    seen_keys = {
        link_import_duplicate_key({"name": link.name, "url": link.url})
        for link in db.scalars(select(models.QuickLink)).all()
    }
    for index, row in enumerate(records, start=2):
        try:
            candidate = normalize_link_import_row(row)
        except (ValueError, ValidationError) as exc:
            skipped_invalid += 1
            errors.append(f"Row {index}: {exc}")
            continue
        duplicate_key = link_import_duplicate_key(candidate)
        if duplicate_key in seen_keys:
            skipped_duplicates += 1
            continue
        db.add(models.QuickLink(**candidate))
        seen_keys.add(duplicate_key)
        imported += 1
    commit_or_conflict(db)
    result = {
        "filename": filename,
        "rows": len(records),
        "imported": imported,
        "skipped_duplicates": skipped_duplicates,
        "skipped_invalid": skipped_invalid,
        "errors": errors[:10],
        "accepted_environments": sorted(LINK_ENVIRONMENTS),
    }
    logger.info(
        "Imported links file=%s rows=%s imported=%s skipped_duplicates=%s skipped_invalid=%s",
        filename,
        result["rows"],
        imported,
        skipped_duplicates,
        skipped_invalid,
    )
    return result


for route_path, route_cfg in MODEL_CONFIG.items():
    register_crud(route_path, route_cfg)


@app.post("/api/reminders/{item_id}/complete")
def complete_reminder(item_id: int, db: Session = Depends(get_db)):
    reminder = get_or_404(db, models.Reminder, item_id)
    reminder.status = "completed"
    db.commit()
    db.refresh(reminder)
    return schemas.ReminderOut.model_validate(reminder)


@app.post("/api/chores/{item_id}/complete")
def complete_chore(item_id: int, db: Session = Depends(get_db)):
    chore = get_or_404(db, models.Chore, item_id)
    if chore.recurrence == "one-time":
        chore.status = "completed"
    else:
        chore.due_date = next_due(chore.due_date, chore.recurrence)
        chore.status = "open"
    db.commit()
    db.refresh(chore)
    return schemas.ChoreOut.model_validate(chore)


@app.post("/api/project-tasks/{item_id}/complete")
def complete_project_task(item_id: int, db: Session = Depends(get_db)):
    task = get_or_404(db, models.ProjectTask, item_id)
    task.status = "completed"
    db.commit()
    db.refresh(task)
    return schemas.ProjectTaskOut.model_validate(task)


@app.post("/api/projects/{item_id}/mark-worked")
def mark_project_worked(item_id: int, db: Session = Depends(get_db)):
    project = get_or_404(db, models.Project, item_id)
    project.last_worked_at = date.today().isoformat()
    commit_or_conflict(db)
    db.refresh(project)
    return schemas.ProjectOut.model_validate(project)


@app.post("/api/projects/{item_id}/archive")
def archive_project(item_id: int, db: Session = Depends(get_db)):
    project = get_or_404(db, models.Project, item_id)
    project.status = "archived"
    commit_or_conflict(db)
    db.refresh(project)
    logger.info("Archived project id=%s", item_id)
    return schemas.ProjectOut.model_validate(project)


@app.post("/api/projects/{item_id}/next-action")
def update_project_next_action(item_id: int, payload: schemas.ProjectNextActionUpdate, db: Session = Depends(get_db)):
    project = get_or_404(db, models.Project, item_id)
    project.next_action = payload.next_action
    project.next_step = payload.next_action
    commit_or_conflict(db)
    db.refresh(project)
    return schemas.ProjectOut.model_validate(project)


@app.post("/api/links/{item_id}/opened")
def mark_link_opened(item_id: int, db: Session = Depends(get_db)):
    link = get_or_404(db, models.QuickLink, item_id)
    link.open_count = (link.open_count or 0) + 1
    link.last_opened_at = now_stamp()
    db.commit()
    db.refresh(link)
    return schemas.QuickLinkOut.model_validate(link)


@app.get("/api/tools/local-system-info")
def local_system_info():
    memory = psutil.virtual_memory()
    boot = datetime.fromtimestamp(psutil.boot_time())
    return {
        "hostname": platform.node(),
        "platform": platform.platform(),
        "os": platform.system(),
        "os_version": platform.version(),
        "uptime_seconds": int(time.time() - psutil.boot_time()),
        "boot_time": boot.isoformat(timespec="seconds"),
        "cpu": {
            "logical_count": psutil.cpu_count(logical=True),
            "physical_count": psutil.cpu_count(logical=False),
            "percent": psutil.cpu_percent(interval=0.1),
            "processor": platform.processor(),
        },
        "memory": {
            "total_gb": round(memory.total / (1024**3), 2),
            "available_gb": round(memory.available / (1024**3), 2),
            "used_gb": round(memory.used / (1024**3), 2),
            "percent": round(memory.percent, 1),
        },
        "storage": storage_metrics(),
        "checked_at": now_stamp(),
    }


@app.post("/api/tools/speed-test")
def speed_test(payload: schemas.SpeedTestRequest, db: Session = Depends(get_db)):
    test_url = payload.test_url or "https://speed.cloudflare.com/__down?bytes=5000000"
    upload_test_url = payload.upload_test_url or "https://speed.cloudflare.com/__up"
    samples = []
    download_mbps = None
    upload_mbps = None
    bytes_read = 0
    bytes_uploaded = 0
    upload_note = ""
    try:
        samples = latency_samples(test_url, count=3)
        download_mbps, bytes_read, _ = measure_download_mbps(test_url)
    except Exception as exc:  # noqa: BLE001 - returned as user-visible tool failure
        logger.exception("Speed test failed for download URL %s", test_url)
        raise HTTPException(status_code=502, detail="Speed test failed") from exc
    try:
        upload_mbps, bytes_uploaded, _ = measure_upload_mbps(upload_test_url)
    except Exception as exc:  # noqa: BLE001 - upload can be blocked while download still succeeds
        logger.exception("Speed test upload failed for upload URL %s", upload_test_url)
        upload_note = " Upload test failed."
    latency = round(sum(samples) / len(samples), 2) if samples else None
    jitter = round(max(samples) - min(samples), 2) if len(samples) > 1 else None
    result = models.SpeedTestResult(
        download_mbps=download_mbps,
        upload_mbps=upload_mbps,
        latency_ms=latency,
        jitter_ms=jitter,
        test_url=test_url,
        method="approximate_http_download_upload",
        notes=(
            f"Approximate backend HTTP test using {bytes_read} downloaded bytes and "
            f"{bytes_uploaded} uploaded bytes. Download URL: {test_url}. Upload URL: {upload_test_url}."
            f"{upload_note}"
        ),
        tested_at=now_stamp(),
    )
    db.add(result)
    commit_or_conflict(db)
    db.refresh(result)
    return schemas.SpeedTestResultOut.model_validate(result)


@app.delete("/api/tools/speed-test-history", status_code=204)
def clear_speed_test_history(db: Session = Depends(get_db)):
    for item in db.scalars(select(models.SpeedTestResult)).all():
        db.delete(item)
    db.commit()
    return None


@app.post("/api/tools/ping")
def ping_test(payload: schemas.PingRequest):
    return run_safe_ping(payload.host)


@app.post("/api/tools/dns")
def dns_lookup(payload: schemas.DnsLookupRequest):
    domain = clamp_host(payload.domain)
    record_type = payload.record_type.upper()
    requested_types = ["A", "AAAA", "CNAME", "MX", "TXT", "NS"] if record_type == "ALL" else [record_type]
    rows = []
    errors = []

    def lookup_type(current_type: str):
        type_rows = []
        if current_type in {"A", "AAAA"}:
            family = socket.AF_INET if current_type == "A" else socket.AF_INET6
            seen = set()
            for result in socket.getaddrinfo(domain, None, family, socket.SOCK_STREAM):
                value = result[4][0]
                if value not in seen:
                    type_rows.append({"name": domain, "type": current_type, "value": value})
                    seen.add(value)
            return type_rows
        if dns is None:
            raise RuntimeError("dnspython is not installed")
        answers = dns.resolver.resolve(domain, current_type, lifetime=5)
        for answer in answers:
            type_rows.append({"name": domain, "type": current_type, "value": answer.to_text()})
        return type_rows

    if record_type != "ALL" and record_type not in {"A", "AAAA"} and dns is None:
        raise HTTPException(status_code=501, detail="This runtime only supports A/AAAA lookups without dnspython installed")

    for current_type in requested_types:
        try:
            rows.extend(lookup_type(current_type))
        except Exception as exc:  # noqa: BLE001
            logger.exception("DNS lookup failed for domain %s record type %s", domain, current_type)
            if record_type != "ALL":
                raise HTTPException(status_code=502, detail="DNS lookup failed") from exc
            errors.append({"type": current_type, "message": "DNS lookup failed"})

    grouped = {current_type: [row for row in rows if row["type"] == current_type] for current_type in requested_types}
    return {"domain": domain, "record_type": record_type, "records": rows, "grouped": grouped, "errors": errors, "checked_at": now_stamp()}


@app.post("/api/tools/port-check")
def check_port(payload: schemas.PortCheckRequest):
    return port_check(payload.host, payload.port)


@app.get("/api/tools/local-apps")
def local_apps():
    apps = installed_windows_apps()
    return {"apps": apps, "count": len(apps), "checked_at": now_stamp(), "source": "windows_uninstall_registry" if winreg is not None else "unsupported"}


@app.post("/api/tools/certificate")
def check_certificate(payload: schemas.CertificateCheckRequest):
    domain = clamp_host(payload.domain)
    try:
        context = ssl.create_default_context()
        with socket.create_connection((domain, payload.port), timeout=6) as sock:
            with context.wrap_socket(sock, server_hostname=domain) as tls:
                cert = tls.getpeercert()
    except Exception as exc:  # noqa: BLE001
        logger.exception("Certificate check failed for domain %s port %s", domain, payload.port)
        raise HTTPException(status_code=502, detail="Certificate check failed") from exc
    not_before = datetime.strptime(cert["notBefore"], "%b %d %H:%M:%S %Y %Z")
    not_after = datetime.strptime(cert["notAfter"], "%b %d %H:%M:%S %Y %Z")
    days_remaining = (not_after.date() - date.today()).days
    issuer_parts = cert.get("issuer", [])
    issuer = ", ".join("=".join(attr) for part in issuer_parts for attr in part)
    return {
        "domain": domain,
        "issuer": issuer,
        "valid_from": not_before.date().isoformat(),
        "valid_to": not_after.date().isoformat(),
        "days_remaining": days_remaining,
        "status": certificate_status(days_remaining),
        "checked_at": now_stamp(),
    }


def convert_inbox_item(item: models.InboxItem, payload: schemas.InboxConvertRequest, db: Session) -> dict[str, Any]:
    raw = item.raw_text.strip()
    if not raw:
        raise HTTPException(status_code=422, detail="Inbox item text cannot be blank")
    if payload.target_type == "reminder":
        created = models.Reminder(title=raw[:160], due_date=payload.due_date, priority=item.priority, notes=raw, inbox_item_id=item.id)
        out_schema = schemas.ReminderOut
    elif payload.target_type == "note":
        note_type = "note"
        created = models.Note(title=raw.splitlines()[0][:160], body=raw, note_type=note_type, tags=item.tags, inbox_item_id=item.id)
        out_schema = schemas.NoteOut
    elif payload.target_type == "project task":
        created = models.ProjectTask(title=raw[:160], project_name=payload.project_name, priority=item.priority, notes=raw, inbox_item_id=item.id)
        out_schema = schemas.ProjectTaskOut
    elif payload.target_type == "calendar item":
        created = models.CalendarItem(title=raw[:160], date=payload.date or date.today().isoformat(), notes=raw, inbox_item_id=item.id)
        out_schema = schemas.CalendarOut
    elif payload.target_type == "link":
        url = payload.url or first_url(raw)
        if not url:
            raise HTTPException(status_code=422, detail="A URL is required to convert this inbox item to a link")
        created = models.QuickLink(name=raw.replace(url, "").strip()[:120] or url[:120], url=url, notes=raw, tags=item.tags, inbox_item_id=item.id)
        out_schema = schemas.QuickLinkOut
    else:
        raise HTTPException(status_code=422, detail="Unsupported conversion target")
    db.add(created)
    item.status = "processed"
    item.processed_at = now_stamp()
    commit_or_conflict(db)
    db.refresh(created)
    db.refresh(item)
    logger.info("Converted inbox id=%s to %s id=%s", item.id, payload.target_type, created.id)
    return {"inbox_item": schemas.InboxOut.model_validate(item), "created": out_schema.model_validate(created)}


@app.post("/api/inbox/{item_id}/convert")
def convert_inbox(item_id: int, payload: schemas.InboxConvertRequest, db: Session = Depends(get_db)):
    return convert_inbox_item(get_or_404(db, models.InboxItem, item_id), payload, db)


@app.post("/api/notes/{item_id}/convert")
def convert_note(item_id: int, payload: schemas.InboxConvertRequest, db: Session = Depends(get_db)):
    note = get_or_404(db, models.Note, item_id)
    inbox = models.InboxItem(
        raw_text=f"{note.title}\n\n{note.body}".strip(),
        suggested_type=payload.target_type,
        status="new",
        priority="medium",
        tags=note.tags,
    )
    db.add(inbox)
    db.flush()
    return convert_inbox_item(inbox, payload, db)


@app.post("/api/daily-plans/start")
def start_day(db: Session = Depends(get_db)):
    today = date.today().isoformat()
    plan = db.scalars(select(models.DailyPlan).where(models.DailyPlan.date == today)).first()
    if plan is None:
        active_project = db.scalars(select(models.Project).where(models.Project.status == "active").order_by(models.Project.priority.desc(), models.Project.updated_at.desc())).first()
        open_reminder = db.scalars(select(models.Reminder).where(models.Reminder.status == "open").order_by(models.Reminder.due_date.is_(None), models.Reminder.due_date)).first()
        inbox_item = db.scalars(select(models.InboxItem).where(models.InboxItem.status == "new").order_by(models.InboxItem.created_at)).first()
        plan = models.DailyPlan(
            date=today,
            main_focus=active_project.name if active_project else None,
            top_action_1=(active_project.next_action or active_project.next_step) if active_project else None,
            top_action_2=open_reminder.title if open_reminder else None,
            top_action_3=inbox_item.raw_text[:300] if inbox_item else None,
        )
        db.add(plan)
        commit_or_conflict(db)
        db.refresh(plan)
    return {
        "plan": schemas.DailyPlanOut.model_validate(plan),
        "calendar": [calendar_to_agenda_item(i) for i in db.scalars(select(models.CalendarItem).where(models.CalendarItem.date == today).order_by(models.CalendarItem.start_time)).all()],
        "due_reminders": serialize(db.scalars(select(models.Reminder).where(models.Reminder.status == "open", models.Reminder.due_date <= today).order_by(models.Reminder.due_date)).all(), schemas.ReminderOut),
        "project_actions": [project_display(p) for p in db.scalars(select(models.Project).where(models.Project.status.in_(["active", "inactive", "planned", "blocked"])).order_by(models.Project.updated_at.desc()).limit(10)).all()],
        "inbox": serialize(db.scalars(select(models.InboxItem).where(models.InboxItem.status == "new").order_by(models.InboxItem.created_at).limit(10)).all(), schemas.InboxOut),
    }


@app.post("/api/daily-plans/end")
def end_day(payload: schemas.DailyPlanEndRequest, db: Session = Depends(get_db)):
    today = date.today().isoformat()
    plan = db.scalars(select(models.DailyPlan).where(models.DailyPlan.date == today)).first()
    if plan is None:
        plan = models.DailyPlan(date=today)
        db.add(plan)
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(plan, key, value)
    plan.completed_at = now_stamp()
    commit_or_conflict(db)
    db.refresh(plan)
    return schemas.DailyPlanOut.model_validate(plan)


@app.get("/api/settings")
def get_settings(db: Session = Depends(get_db)):
    settings = db.scalars(select(models.Settings).order_by(models.Settings.id)).first()
    if settings is None:
        settings = models.Settings()
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return schemas.SettingsOut.model_validate(settings)


@app.put("/api/settings")
def update_settings(payload: schemas.SettingsBase, db: Session = Depends(get_db)):
    settings = db.scalars(select(models.Settings).order_by(models.Settings.id)).first()
    if settings is None:
        settings = models.Settings(**payload.model_dump())
        db.add(settings)
    else:
        for key, value in payload.model_dump().items():
            setattr(settings, key, value)
    db.commit()
    db.refresh(settings)
    return schemas.SettingsOut.model_validate(settings)


@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db)):
    today = date.today()
    tomorrow = today + timedelta(days=1)
    in_7 = today + timedelta(days=7)
    in_14 = today + timedelta(days=14)
    in_30 = today + timedelta(days=30)

    auto_complete_past_calendar_reminders(db, commit=True)
    agenda = combined_agenda(db, today, in_7)
    today_agenda = [item for item in agenda if item["date"] == today.isoformat()]
    today_calendar = [item for item in today_agenda if item.get("source") != "reminder"]
    reminders = db.scalars(select(models.Reminder).where(models.Reminder.status == "open").order_by(models.Reminder.due_date, models.Reminder.due_time)).all()
    bills = db.scalars(select(models.Bill).where(models.Bill.status.in_(["active", "review"])).order_by(models.Bill.due_date)).all()
    chores = db.scalars(select(models.Chore).where(models.Chore.status == "open").order_by(models.Chore.due_date)).all()
    projects = db.scalars(select(models.Project).order_by(models.Project.due_date.is_(None), models.Project.due_date, models.Project.name)).all()
    project_tasks = db.scalars(select(models.ProjectTask).where(models.ProjectTask.status == "open").order_by(models.ProjectTask.due_date.is_(None), models.ProjectTask.due_date, models.ProjectTask.due_time)).all()
    notes = db.scalars(select(models.Note).where(models.Note.pinned.is_(True)).order_by(models.Note.updated_at.desc())).all()
    links = db.scalars(select(models.QuickLink).order_by(models.QuickLink.favorite.desc(), models.QuickLink.last_opened_at.desc().nullslast(), models.QuickLink.category, models.QuickLink.name)).all()
    inbox = db.scalars(select(models.InboxItem).where(models.InboxItem.status == "new").order_by(models.InboxItem.created_at.desc()).limit(8)).all()
    daily_plan = db.scalars(select(models.DailyPlan).where(models.DailyPlan.date == today.isoformat())).first()
    recent_speed_test = db.scalars(select(models.SpeedTestResult).order_by(models.SpeedTestResult.tested_at.desc()).limit(1)).first()
    assets_needing_attention = db.scalars(
        select(models.Asset).where(models.Asset.status.in_(["needs_attention", "offline", "unknown"])).order_by(models.Asset.updated_at.desc()).limit(6)
    ).all()
    security_due = db.scalars(
        select(models.SecurityRecord)
        .where(or_(models.SecurityRecord.next_review_date <= in_14.isoformat(), models.SecurityRecord.status.in_(["review_needed", "expiring_soon", "expired"])))
        .order_by(models.SecurityRecord.next_review_date.is_(None), models.SecurityRecord.next_review_date)
        .limit(6)
    ).all()
    expiring_security = db.scalars(
        select(models.SecurityRecord)
        .where(models.SecurityRecord.expiration_date.is_not(None), models.SecurityRecord.expiration_date <= in_30.isoformat())
        .order_by(models.SecurityRecord.expiration_date)
        .limit(6)
    ).all()
    stale_cutoff = (today - timedelta(days=7)).isoformat()
    active_projects = [p for p in projects if schemas.normalize_project_status(p.status) == "active"]
    current_projects = [p for p in active_projects if p.last_worked_at and p.last_worked_at >= stale_cutoff]
    inactive_projects = [
        p for p in projects
        if schemas.normalize_project_status(p.status) == "inactive"
        or (schemas.normalize_project_status(p.status) == "active" and (not p.last_worked_at or p.last_worked_at < stale_cutoff))
    ]

    bill_total = sum(b.amount for b in bills if b.billing_cycle == "monthly")
    return {
        "date": today.isoformat(),
        "today": {
            "calendar": today_calendar,
            "agenda": today_agenda,
            "reminders": serialize([r for r in reminders if r.due_date == today.isoformat()], schemas.ReminderOut),
            "chores": serialize([c for c in chores if c.due_date == today.isoformat()], schemas.ChoreOut),
        },
        "tomorrow": [item for item in agenda if item["date"] == tomorrow.isoformat()],
        "next_7_days": agenda,
        "reminders": {
            "overdue": serialize([r for r in reminders if r.due_date and r.due_date < today.isoformat()], schemas.ReminderOut),
            "upcoming": serialize([r for r in reminders if r.due_date and today.isoformat() < r.due_date <= in_7.isoformat()], schemas.ReminderOut),
        },
        "bills": {
            "monthly_total": round(bill_total, 2),
            "next_7": serialize([b for b in bills if b.due_date <= in_7.isoformat()], schemas.BillOut),
            "next_14": serialize([b for b in bills if b.due_date <= in_14.isoformat()], schemas.BillOut),
            "next_30": serialize([b for b in bills if b.due_date <= in_30.isoformat()], schemas.BillOut),
        },
        "chores": serialize(chores[:10], schemas.ChoreOut),
        "projects": [project_display(p) for p in projects[:8]],
        "project_momentum": {
            "active": [project_display(p) for p in current_projects[:8]],
            "inactive": [project_display(p) for p in inactive_projects[:8]],
            "next_actions": [project_display(p) for p in current_projects if (p.next_action or p.next_step)][:8],
        },
        "project_tasks": serialize(project_tasks[:8], schemas.ProjectTaskOut),
        "notes": serialize(notes[:6], schemas.NoteOut),
        "links": [link_display(link) for link in links],
        "favorite_links": [link_display(link) for link in links if link.favorite][:8],
        "recent_links": [link_display(link) for link in links if link.last_opened_at][:8],
        "inbox": serialize(inbox, schemas.InboxOut),
        "daily_plan": schemas.DailyPlanOut.model_validate(daily_plan).model_dump() if daily_plan else None,
        "tools": {
            "recent_speed_test": schemas.SpeedTestResultOut.model_validate(recent_speed_test).model_dump() if recent_speed_test else None,
            "assets_needing_attention": [asset_display(asset) for asset in assets_needing_attention],
            "security_reviews_due": [security_record_display(record) for record in security_due],
            "expiring_security": [security_record_display(record) for record in expiring_security],
        },
    }


@app.get("/api/agenda")
def agenda(days_back: int = Query(31, ge=0, le=365), days_forward: int = Query(90, ge=1, le=730), db: Session = Depends(get_db)):
    today = date.today()
    return combined_agenda(db, today - timedelta(days=days_back), today + timedelta(days=days_forward))


@app.post("/api/calendar/reminders/backfill")
def calendar_reminders_backfill(db: Session = Depends(get_db)):
    return backfill_calendar_reminders(db)


@app.get("/api/integrations/google-calendar/status")
def google_calendar_integration_status():
    return google_calendar_status()


@app.post("/api/integrations/google-calendar/connect")
def google_calendar_connect():
    return connect_google_calendar()


@app.post("/api/integrations/google-calendar/sync")
def google_calendar_sync(payload: schemas.GoogleCalendarSyncRequest, db: Session = Depends(get_db)):
    return sync_google_calendar(db, payload.calendar_id, payload.days_back, payload.days_forward)


@app.get("/api/integrations/google/config")
def google_oauth_integration_config():
    return google_oauth_config_status()


@app.put("/api/integrations/google/config")
def google_oauth_integration_config_update(payload: schemas.GoogleOAuthConfigUpdate):
    return save_google_oauth_config(payload.client_secret_json)


@app.get("/api/integrations/apple-calendar/status")
def apple_calendar_integration_status():
    return apple_calendar_status()


@app.get("/api/integrations/apple-calendar/config")
def apple_calendar_integration_config():
    return apple_calendar_config()


@app.put("/api/integrations/apple-calendar/config")
def apple_calendar_integration_config_update(payload: schemas.AppleCalendarConfigUpdate):
    return save_apple_calendar_config(payload)


@app.get("/api/integrations/apple-calendar/calendars")
def apple_calendar_calendars():
    return list_apple_calendars()


@app.post("/api/integrations/apple-calendar/sync")
def apple_calendar_sync(payload: schemas.AppleCalendarSyncRequest, db: Session = Depends(get_db)):
    return sync_apple_calendar(db, payload.calendar_name, payload.days_back, payload.days_forward)


@app.get("/api/integrations/gmail/status")
def gmail_integration_status():
    return gmail_status()


@app.post("/api/integrations/gmail/connect")
def gmail_connect():
    return connect_gmail()


@app.post("/api/integrations/gmail/scan-bills")
def gmail_scan_bills(payload: schemas.GmailBillScanRequest, db: Session = Depends(get_db)):
    return scan_gmail_for_bills(db, payload.query, payload.max_messages, payload.keyword_filter)


@app.get("/api/search")
def search(q: str = Query(min_length=1), db: Session = Depends(get_db)):
    needle = f"%{q}%"
    results: list[dict[str, Any]] = []
    search_defs = [
        ("calendar", models.CalendarItem, [models.CalendarItem.title, models.CalendarItem.location, models.CalendarItem.notes, models.CalendarItem.category], "title", "date", "notes", "status", "/calendar"),
        ("reminder", models.Reminder, [models.Reminder.title, models.Reminder.notes, models.Reminder.priority], "title", "due_date", "notes", "status", "/reminders"),
        ("bill", models.Bill, [models.Bill.name, models.Bill.category, models.Bill.billing_cycle, models.Bill.notes, models.Bill.status, models.Bill.source, models.Bill.url, cast(models.Bill.amount, String)], "name", "due_date", "notes", "status", "/bills"),
        ("knowledge_base", models.Note, [models.Note.title, models.Note.body, models.Note.tags, models.Note.note_type], "title", "updated_at", "body", None, "/notes"),
        ("asset", models.Asset, [models.Asset.name, models.Asset.type, models.Asset.role, models.Asset.hostname, models.Asset.ip_address, models.Asset.platform, models.Asset.environment, models.Asset.status, models.Asset.notes], "name", "updated_at", "notes", "status", "/systems"),
        ("security_record", models.SecurityRecord, [models.SecurityRecord.name, models.SecurityRecord.type, models.SecurityRecord.provider, models.SecurityRecord.status, models.SecurityRecord.risk_level, models.SecurityRecord.storage_reference, models.SecurityRecord.notes], "name", "next_review_date", "notes", "status", "/security"),
        ("speed_test", models.SpeedTestResult, [models.SpeedTestResult.test_url, models.SpeedTestResult.method, models.SpeedTestResult.notes], "method", "tested_at", "notes", None, "/systems"),
        ("link", models.QuickLink, [models.QuickLink.name, models.QuickLink.category, models.QuickLink.notes, models.QuickLink.url, models.QuickLink.tags], "name", "updated_at", "url", None, "/links"),
        ("project", models.Project, [models.Project.name, models.Project.status, models.Project.priority, models.Project.category, models.Project.local_path, models.Project.repo_url, models.Project.goal, models.Project.next_action, models.Project.blocker, models.Project.codex_prompt, models.Project.notes, models.Project.tags], "name", "due_date", "next_action", "status", "/projects"),
        ("project_task", models.ProjectTask, [models.ProjectTask.title, models.ProjectTask.project_name, models.ProjectTask.priority, models.ProjectTask.status, models.ProjectTask.codex_prompt, models.ProjectTask.notes], "title", "due_date", "notes", "status", "/projects"),
        ("inbox", models.InboxItem, [models.InboxItem.raw_text, models.InboxItem.suggested_type, models.InboxItem.tags], "raw_text", "created_at", "raw_text", "status", "/inbox"),
        ("daily_plan", models.DailyPlan, [models.DailyPlan.date, models.DailyPlan.main_focus, models.DailyPlan.top_action_1, models.DailyPlan.top_action_2, models.DailyPlan.top_action_3, models.DailyPlan.end_of_day_note, models.DailyPlan.tomorrow_first_action], "main_focus", "date", "end_of_day_note", None, "/"),
    ]
    for item_type, model, fields, title_field, date_field, snippet_field, status_field, target in search_defs:
        stmt = select(model).where(or_(*[field.ilike(needle) for field in fields])).limit(25)
        for item in db.scalars(stmt).all():
            title = getattr(item, title_field) or ""
            snippet = getattr(item, snippet_field, None) if snippet_field else None
            status = getattr(item, status_field, None) if status_field else None
            if item_type == "link":
                target = getattr(item, "url")
            results.append(
                {
                    "type": item_type,
                    "id": item.id,
                    "title": title[:160] or item_type,
                    "snippet": (snippet or "")[:240],
                    "relevant_date": getattr(item, date_field),
                    "status": status,
                    "action_target": target,
                    "copy_link": getattr(item, "url", None) if item_type == "link" else None,
                    "can_attach_to_project": item_type in {"knowledge_base", "link", "inbox", "daily_plan", "asset", "security_record"},
                }
            )
    return sorted(results, key=lambda x: (x["relevant_date"] or "9999-99-99", x["type"], x["title"]))[:100]


@app.post("/api/search/actions/create-reminder")
def search_create_reminder(payload: schemas.SearchActionRequest, db: Session = Depends(get_db)):
    model_map = {
        "calendar": models.CalendarItem,
        "bill": models.Bill,
        "knowledge_base": models.Note,
        "link": models.QuickLink,
        "project": models.Project,
        "project_task": models.ProjectTask,
        "inbox": models.InboxItem,
        "daily_plan": models.DailyPlan,
        "asset": models.Asset,
        "security_record": models.SecurityRecord,
        "speed_test": models.SpeedTestResult,
    }
    model = model_map.get(payload.result_type)
    if model is None:
        raise HTTPException(status_code=422, detail="Unsupported search result type")
    item = get_or_404(db, model, payload.result_id)
    title = getattr(item, "title", None) or getattr(item, "name", None) or getattr(item, "raw_text", None) or getattr(item, "main_focus", None) or payload.result_type
    reminder = models.Reminder(title=str(title)[:160], notes=f"Created from search result {payload.result_type} #{payload.result_id}", priority="medium")
    db.add(reminder)
    commit_or_conflict(db)
    db.refresh(reminder)
    return schemas.ReminderOut.model_validate(reminder)


@app.post("/api/search/actions/attach-to-project")
def search_attach_to_project(payload: schemas.SearchActionRequest, db: Session = Depends(get_db)):
    if payload.project_id is None:
        raise HTTPException(status_code=422, detail="project_id is required")
    project = get_or_404(db, models.Project, payload.project_id)
    if payload.result_type not in {"knowledge_base", "link", "inbox", "daily_plan", "asset", "security_record"}:
        raise HTTPException(status_code=422, detail="This result type cannot be attached to a project")
    task = models.ProjectTask(
        title=f"Review {payload.result_type} #{payload.result_id}",
        project_name=project.name,
        notes=f"Attached from search result {payload.result_type} #{payload.result_id}",
        priority="medium",
    )
    db.add(task)
    commit_or_conflict(db)
    db.refresh(task)
    return schemas.ProjectTaskOut.model_validate(task)


@app.get("/api/ai-helper/{prompt_key}")
def ai_helper(prompt_key: str, db: Session = Depends(get_db)):
    data = dashboard(db)
    lines: list[str] = []
    if prompt_key == "attention-today":
        lines.append(f"{len(data['reminders']['overdue'])} overdue reminder(s).")
        lines.append(f"{len(data['today']['agenda'])} agenda item(s) today.")
        lines.append(f"{len(data['today']['chores'])} chore(s) due today.")
        lines.append(f"{len(data['bills']['next_7'])} bill(s) due in the next 7 days.")
    elif prompt_key == "overdue":
        overdue = data["reminders"]["overdue"]
        lines = [f"{item['title']} was due {item['due_date']}." for item in overdue] or ["No overdue reminders."]
    elif prompt_key == "bills-coming-up":
        bills = data["bills"]["next_30"]
        lines = [f"{b['name']}: ${b['amount']:.2f} due {b['due_date']} ({b['billing_cycle']})." for b in bills] or ["No bills due in the next 30 days."]
    elif prompt_key == "summarize-week":
        lines.append(f"This week has {len(data['next_7_days'])} agenda item(s), {len(data['reminders']['upcoming'])} upcoming reminder(s), and {len(data['chores'])} active chore(s).")
        if data["bills"]["next_7"]:
            total = sum(b["amount"] for b in data["bills"]["next_7"])
            lines.append(f"Bills due this week total ${total:.2f}.")
    else:
        raise HTTPException(status_code=404, detail="Unknown helper prompt")
    return {"provider": "rules", "summary": " ".join(lines), "bullets": lines}


@app.get("/api/export")
def export_data(db: Session = Depends(get_db)):
    return {
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "calendar_items": serialize(db.scalars(select(models.CalendarItem)).all(), schemas.CalendarOut),
        "reminders": serialize(db.scalars(select(models.Reminder)).all(), schemas.ReminderOut),
        "bills": serialize(db.scalars(select(models.Bill)).all(), schemas.BillOut),
        "chores": serialize(db.scalars(select(models.Chore)).all(), schemas.ChoreOut),
        "notes": serialize(db.scalars(select(models.Note)).all(), schemas.NoteOut),
        "assets": serialize(db.scalars(select(models.Asset)).all(), schemas.AssetOut),
        "speed_test_results": serialize(db.scalars(select(models.SpeedTestResult)).all(), schemas.SpeedTestResultOut),
        "security_records": serialize(db.scalars(select(models.SecurityRecord)).all(), schemas.SecurityRecordOut),
        "quick_links": serialize(db.scalars(select(models.QuickLink)).all(), schemas.QuickLinkOut),
        "projects": serialize(db.scalars(select(models.Project)).all(), schemas.ProjectOut),
        "project_tasks": serialize(db.scalars(select(models.ProjectTask)).all(), schemas.ProjectTaskOut),
        "inbox_items": serialize(db.scalars(select(models.InboxItem)).all(), schemas.InboxOut),
        "daily_plans": serialize(db.scalars(select(models.DailyPlan)).all(), schemas.DailyPlanOut),
        "settings": get_settings(db),
    }


@app.post("/api/import")
def import_data(payload: schemas.ImportPayload, db: Session = Depends(get_db)):
    mapping = [
        ("calendar_items", models.CalendarItem),
        ("reminders", models.Reminder),
        ("bills", models.Bill),
        ("chores", models.Chore),
        ("notes", models.Note),
        ("assets", models.Asset),
        ("speed_test_results", models.SpeedTestResult),
        ("security_records", models.SecurityRecord),
        ("quick_links", models.QuickLink),
        ("projects", models.Project),
        ("project_tasks", models.ProjectTask),
        ("inbox_items", models.InboxItem),
        ("daily_plans", models.DailyPlan),
    ]
    for attr, model in mapping:
        for item in getattr(payload, attr):
            db.add(model(**item.model_dump()))
    if payload.settings:
        update_settings(payload.settings, db)
    commit_or_conflict(db)
    return {"status": "imported"}


@app.get("/api/export/bills.csv")
def export_bills_csv(db: Session = Depends(get_db)):
    rows = db.scalars(select(models.Bill).order_by(models.Bill.due_date)).all()
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["name", "amount", "due_date", "billing_cycle", "category", "autopay", "url", "status", "notes"])
    for bill in rows:
        writer.writerow([bill.name, bill.amount, bill.due_date, bill.billing_cycle, bill.category, bill.autopay, bill.url, bill.status, bill.notes])
    return StreamingResponse(iter([buffer.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=bills.csv"})


@app.get("/api/export/bills.xlsx")
def export_bills_xlsx(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise HTTPException(status_code=500, detail="XLSX export requires openpyxl. Install backend requirements first.") from exc
    rows = db.scalars(select(models.Bill).order_by(models.Bill.due_date)).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bills"
    headers = ["name", "amount", "due_date", "billing_cycle", "category", "autopay", "url", "status", "notes"]
    sheet.append(headers)
    for bill in rows:
        sheet.append([bill.name, bill.amount, bill.due_date, bill.billing_cycle, bill.category, bill.autopay, bill.url, bill.status, bill.notes])
    for cell in sheet[1]:
        cell.style = "Headline 3"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bills.xlsx"},
    )


@app.get("/api/export/notes.csv")
def export_notes_csv(db: Session = Depends(get_db)):
    rows = db.scalars(select(models.Note).order_by(models.Note.title)).all()
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Title", "Category", "Tags", "Pinned", "Body"])
    for note in rows:
        writer.writerow([note.title, note.note_type, note.tags or "", "TRUE" if note.pinned else "FALSE", note.body or ""])
    return StreamingResponse(iter([buffer.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=knowledge-base-articles.csv"})


@app.get("/api/export/notes.xlsx")
def export_notes_xlsx(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise HTTPException(status_code=500, detail="XLSX export requires openpyxl. Install backend requirements first.") from exc
    rows = db.scalars(select(models.Note).order_by(models.Note.title)).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Knowledge Base"
    sheet.append(["Title", "Category", "Tags", "Pinned", "Body"])
    for note in rows:
        sheet.append([note.title, note.note_type, note.tags or "", "TRUE" if note.pinned else "FALSE", note.body or ""])
    for cell in sheet[1]:
        cell.style = "Headline 3"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 64)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=knowledge-base-articles.xlsx"},
    )


@app.get("/api/export/links.csv")
def export_links_csv(db: Session = Depends(get_db)):
    rows = db.scalars(select(models.QuickLink).order_by(models.QuickLink.name)).all()
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Name", "URL", "Category", "Environment", "Tags", "Favorite", "Body"])
    for link in rows:
        writer.writerow([
            link.name,
            link.url,
            link.category or "",
            link.environment or "personal",
            link.tags or "",
            "TRUE" if link.favorite else "FALSE",
            link.notes or "",
        ])
    return StreamingResponse(iter([buffer.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=knowledge-base-links.csv"})


@app.get("/api/export/links.xlsx")
def export_links_xlsx(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise HTTPException(status_code=500, detail="XLSX export requires openpyxl. Install backend requirements first.") from exc
    rows = db.scalars(select(models.QuickLink).order_by(models.QuickLink.name)).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Links"
    sheet.append(["Name", "URL", "Category", "Environment", "Tags", "Favorite", "Body"])
    for link in rows:
        sheet.append([
            link.name,
            link.url,
            link.category or "",
            link.environment or "personal",
            link.tags or "",
            "TRUE" if link.favorite else "FALSE",
            link.notes or "",
        ])
    for cell in sheet[1]:
        cell.style = "Headline 3"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 64)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=knowledge-base-links.xlsx"},
    )


@app.get("/api/export/notes.md")
def export_notes_markdown(db: Session = Depends(get_db)):
    notes = db.scalars(select(models.Note).order_by(models.Note.title)).all()
    content = "\n\n".join([f"# {n.title}\n\nTags: {n.tags or ''}\n\n{n.body}" for n in notes])
    return PlainTextResponse(content, media_type="text/markdown")
